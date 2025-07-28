#!/usr/bin/env python3
"""
Excel to JSON Converter - Streamlit Application

This Streamlit app allows users to upload an Excel file (Passport) and converts
it to JSON format suitable for API input. Multiple rows are treated as one deal
with multiple buildings and services.
"""

import streamlit as st
import pandas as pd
import json
import io
from openpyxl import load_workbook
from datetime import datetime
import traceback

class ExcelToJSONConverter:
    """Convert Excel Passport data to API JSON format"""
    
    def __init__(self):
        self.buildings_data = []
        self.customer_info = {}
        
    def load_excel_data(self, uploaded_file):
        """Load and parse Excel data from uploaded file"""
        try:
            # Load workbook from uploaded file
            wb = load_workbook(uploaded_file, data_only=True)
            
            # Check if 'Janitorial Services' sheet exists
            if 'Janitorial Services' not in wb.sheetnames:
                st.error("Excel file must contain a 'Janitorial Services' sheet")
                return False
                
            ws = wb['Janitorial Services']
            
            # Column mapping (1-based Excel columns)
            column_map = {
                'customer': 'B',           # CUSTOMER
                'building_id': 'C',        # BUILDING ID  
                'address': 'D',            # ADDRESS
                'city': 'E',
                'state': 'F',              # STATE
                'zip': 'G',                # ZIP
                'building_type': 'I',      # BUILDING TYPE
                'total_sq_ft': 'J',        # TOTAL SQ FOOTAGE
                'cleanable_sq_ft': 'K',    # CLEANABLE SQ FOOTAGE
                'alternate_productivity': 'M', # ALTERNATE PRODUCTIVITY
                
                # Schedule columns
                'sun': 'R', 'mon': 'S', 'tue': 'T', 'wed': 'U', 
                'thu': 'V', 'fri': 'W', 'sat': 'X',
                
                # Additional costs
                'additional_costs': 'AB',  # ADDITIONAL COSTS
                
                # Equipment columns
                'equipment_rental_1': 'AH',  # EQUIPMENT RENTAL #1
                'contract_terms_1': 'AI',    # CONTRACT #1 TERMS
                'equipment_rental_2': 'AJ',  # EQUIPMENT RENTAL #2
                'contract_terms_2': 'AK',    # CONTRACT #2 TERMS
                
                # Day Porter columns
                'wage_adjustment': 'AO',     # WAGE ADJUSTMENT
                'dp_sun': 'AQ', 'dp_mon': 'AR', 'dp_tue': 'AS', 'dp_wed': 'AT',
                'dp_thu': 'AU', 'dp_fri': 'AV', 'dp_sat': 'AW',
                
                # Supervisor columns
                'sup_sun': 'AZ', 'sup_mon': 'BA', 'sup_tue': 'BB', 'sup_wed': 'BC',
                'sup_thu': 'BD', 'sup_fri': 'BE', 'sup_sat': 'BF',
            }
            
            self.buildings_data = []
            customer_name = None
            
            # Extract data starting from row 4
            row = 4
            while row <= ws.max_row and row <= 100:  # Limit to prevent infinite loop
                # Check if there's a customer name
                customer_cell = ws[f'B{row}']
                if not customer_cell.value:
                    row += 1
                    continue
                
                # Check if building ID exists (essential field)
                building_id_cell = ws[f'C{row}']
                if not building_id_cell.value or str(building_id_cell.value).strip() == '':
                    row += 1
                    continue
                
                # Check if cleanable square footage exists and is valid
                cleanable_sq_ft_cell = ws[f'K{row}']
                if not cleanable_sq_ft_cell.value or cleanable_sq_ft_cell.value == 0:
                    row += 1
                    continue
                    
                building_data = {}
                
                # Extract all data using column mapping
                for field, col in column_map.items():
                    cell = ws[f'{col}{row}']
                    value = cell.value
                    
                    # Clean up the value - preserve spacing for facility types
                    if value is None:
                        building_data[field] = '' if field in ['customer', 'building_id', 'address', 'city', 'state', 'building_type'] else 0
                    elif isinstance(value, str):
                        # Don't strip building_type (facility type) - preserve exact spacing
                        if field == 'building_type':
                            building_data[field] = value
                        else:
                            building_data[field] = value.strip()
                    else:
                        building_data[field] = value
                
                # Convert numeric fields that might be stored as strings
                numeric_fields = ['cleanable_sq_ft', 'total_sq_ft', 'zip', 'alternate_productivity',
                                'sun', 'mon', 'tue', 'wed', 'thu', 'fri', 'sat',
                                'additional_costs', 'wage_adjustment',
                                'dp_sun', 'dp_mon', 'dp_tue', 'dp_wed', 'dp_thu', 'dp_fri', 'dp_sat',
                                'sup_sun', 'sup_mon', 'sup_tue', 'sup_wed', 'sup_thu', 'sup_fri', 'sup_sat']
                
                for field in numeric_fields:
                    if field in building_data and building_data[field] == '':
                        building_data[field] = 0
                    elif field in building_data and isinstance(building_data[field], str):
                        try:
                            building_data[field] = float(building_data[field])
                        except ValueError:
                            building_data[field] = 0
                
                # Store customer name from first row
                if customer_name is None:
                    customer_name = building_data['customer']
                
                self.buildings_data.append(building_data)
                row += 1
            
            # Store customer info
            if self.buildings_data:
                self.customer_info = {
                    'customer_name': customer_name or "Unknown Customer"
                }
                
            return True
            
        except Exception as e:
            st.error(f"Error loading Excel file: {str(e)}")
            st.error("Please ensure the file is a valid Excel file with the correct format.")
            return False
    
    def convert_to_api_json(self):
        """Convert loaded Excel data to API JSON format"""
        if not self.buildings_data:
            return None
        
        # Get customer name and create record IDs
        customer_name = self.customer_info.get('customer_name', 'Unknown Customer')
        deal_record_id = f"DEAL_{customer_name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
        customer_record_id = f"CUST_{customer_name}"
        
        buildings = []
        
        for i, building_data in enumerate(self.buildings_data):
            # Map contract terms to API format
            contract_terms_map = {
                '12': '12', '24': '24', '36': '36', '60': '60',
                12: '12', 24: '24', 36: '36', 60: '60'
            }
            
            # Equipment mapping
            equipment = []
            if building_data.get('equipment_rental_1'):
                equipment.append({
                    "equipmentType": building_data['equipment_rental_1'],
                    "contractTerm": contract_terms_map.get(building_data.get('contract_terms_1', ''), '')
                })
            if building_data.get('equipment_rental_2'):
                equipment.append({
                    "equipmentType": building_data['equipment_rental_2'],
                    "contractTerm": contract_terms_map.get(building_data.get('contract_terms_2', ''), '')
                })
            
            # Day porter hours
            dayporter_hours = {
                "sunday": building_data.get('dp_sun', 0),
                "monday": building_data.get('dp_mon', 0),
                "tuesday": building_data.get('dp_tue', 0),
                "wednesday": building_data.get('dp_wed', 0),
                "thursday": building_data.get('dp_thu', 0),
                "friday": building_data.get('dp_fri', 0),
                "saturday": building_data.get('dp_sat', 0)
            }
            
            # Supervisor hours
            supervisor_hours = {
                "sunday": building_data.get('sup_sun', 0),
                "monday": building_data.get('sup_mon', 0),
                "tuesday": building_data.get('sup_tue', 0),
                "wednesday": building_data.get('sup_wed', 0),
                "thursday": building_data.get('sup_thu', 0),
                "friday": building_data.get('sup_fri', 0),
                "saturday": building_data.get('sup_sat', 0)
            }
            
            # Productivity override
            productivity_override = {}
            if building_data.get('alternate_productivity'):
                productivity_override = {
                    "value": building_data['alternate_productivity']
                }
            
            # Convert monthly additional costs to weekly for API
            monthly_additional_costs = building_data.get('additional_costs', 0)
            weeks_per_month = 4.33  # Standard assumption
            weekly_additional_costs = monthly_additional_costs / weeks_per_month if weeks_per_month > 0 else 0
            
            # Cost adjustments
            cost_adjustments = {
                "hourlyWageAdjustment": 0,  # Not specified in mapping
                "weeklyAdditionalCosts": weekly_additional_costs
            }
            
            # Create building object
            building = {
                "buildingRecordId": f"BLDG_{building_data['building_id']}",
                "buildingId": str(building_data['building_id']),
                "buildingName": f"{customer_name} - {building_data['building_id']}",
                "facilityType": building_data['building_type'],
                "location": {
                    "state": building_data['state'],
                    "postalCode": str(int(building_data['zip'])) if building_data['zip'] else "",
                    "address": building_data['address'],
                    "city": building_data['city'],
                    "country": "USA",
                },
                "buildingDetails": {
                    "totalSquareFootage": building_data.get('total_sq_ft', 0),
                    "cleanableSquareFootage": building_data['cleanable_sq_ft']
                },
                "services": [
                    {
                        "lineItemObjectId": f"LINE_{building_data['building_id']}",
                        "serviceType": "RJS",
                        "serviceFrequency": "weekly",
                        "schedule": {
                            "sunday": building_data.get('sun', 0),
                            "monday": building_data.get('mon', 0),
                            "tuesday": building_data.get('tue', 0),
                            "wednesday": building_data.get('wed', 0),
                            "thursday": building_data.get('thu', 0),
                            "friday": building_data.get('fri', 0),
                            "saturday": building_data.get('sat', 0)
                        },
                        "inputs": [
                            {
                                "itemName": "cleanableSquareFootage",
                                "itemValue": building_data['cleanable_sq_ft']
                            }
                        ],
                        "productivityOverride": productivity_override,
                        "costAdjustments": cost_adjustments,
                        "equipment": equipment,
                        "dayporterHours": dayporter_hours,
                        "dayporterHourlyWageAdjustment": building_data.get('wage_adjustment', 0),
                        "supervisorHours": supervisor_hours
                    }
                ]
            }
            
            buildings.append(building)
        
        # Create final API JSON structure
        api_json = {
            "dealRecordId": deal_record_id,
            "customerRecordId": customer_record_id,
            "buildings": buildings
        }
        
        return api_json

def main():
    """Main Streamlit application"""
    st.set_page_config(
        page_title="Excel to JSON Converter",
        page_icon="📊",
        layout="wide"
    )
    
    # Initialize session state
    if 'initialized' not in st.session_state:
        st.session_state.initialized = True
    
    st.title("📊 Excel Passport to JSON Converter")
    st.markdown("Convert your Excel Passport file to JSON format for API input")
    
    # File upload section
    st.header("1. Upload Excel File")
    uploaded_file = st.file_uploader(
        "Choose an Excel file (.xlsx or .xlsm)",
        type=['xlsx', 'xlsm'],
        help="Upload your Passport Excel file with Janitorial Services sheet"
    )
    
    if uploaded_file is not None:
        st.success(f"✅ File uploaded: {uploaded_file.name}")
        
        # Initialize converter
        converter = ExcelToJSONConverter()
        
        # Load and process the Excel file
        with st.spinner("Processing Excel file..."):
            if converter.load_excel_data(uploaded_file):
                st.success(f"✅ Successfully loaded {len(converter.buildings_data)} building(s)")
                
                # Display summary
                st.header("2. Data Summary")
                col1, col2 = st.columns(2)
                
                with col1:
                    st.metric("Customer", converter.customer_info.get('customer_name', 'Unknown'))
                    st.metric("Total Buildings", len(converter.buildings_data))
                
                with col2:
                    total_sq_ft = sum(b.get('cleanable_sq_ft', 0) for b in converter.buildings_data)
                    st.metric("Total Cleanable Sq Ft", f"{total_sq_ft:,.0f}")
                
                # Show building details
                if st.checkbox("Show building details"):
                    for i, building in enumerate(converter.buildings_data):
                        with st.expander(f"Building {i+1}: {building.get('building_id', 'Unknown')}"):
                            col1, col2, col3 = st.columns(3)
                            with col1:
                                st.write(f"**Building ID:** {building.get('building_id', 'N/A')}")
                                st.write(f"**Type:** {building.get('building_type', 'N/A')}")
                                st.write(f"**Cleanable Sq Ft:** {building.get('cleanable_sq_ft', 0):,.0f}")
                            with col2:
                                st.write(f"**Address:** {building.get('address', 'N/A')}")
                                st.write(f"**City:** {building.get('city', 'N/A')}")
                                st.write(f"**State:** {building.get('state', 'N/A')}")
                            with col3:
                                schedule = [f"{day}: {building.get(day, 0)}" for day in ['sun', 'mon', 'tue', 'wed', 'thu', 'fri', 'sat'] if building.get(day, 0) > 0]
                                st.write(f"**Schedule:** {', '.join(schedule) if schedule else 'No schedule'}")
                
                # Generate JSON
                st.header("3. Generated JSON")
                
                try:
                    api_json = converter.convert_to_api_json()
                    
                    if api_json:
                        # Display JSON in tabs
                        tab1, tab2, tab3 = st.tabs(["📋 Formatted JSON", "📝 Raw JSON", "📥 Download"])
                        
                        with tab1:
                            st.json(api_json)
                        
                        with tab2:
                            json_string = json.dumps(api_json, indent=2)
                            st.code(json_string, language='json')
                        
                        with tab3:
                            st.subheader("Download & Copy Options")
                            
                            # Prepare JSON string once
                            json_string = json.dumps(api_json, indent=2)
                            
                            # Use file hash for unique key instead of timestamp
                            import hashlib
                            file_hash = hashlib.md5(uploaded_file.name.encode()).hexdigest()[:8]
                            
                            # Generate filename
                            customer_safe = converter.customer_info.get('customer_name', 'Unknown').replace(' ', '_').replace('/', '_')
                            filename = f"{customer_safe}_api_input.json"
                            
                            # Alternative download approach - show the data and let browser handle it
                            col1, col2 = st.columns(2)
                            
                            with col1:
                                st.markdown("**Option 1: Direct Download**")
                                # Simpler download button
                                if st.button("📥 Prepare Download", key=f"prep_download_{file_hash}"):
                                    st.success("✅ JSON prepared! Use the download link below:")
                                    
                                # Always show download button after JSON is ready
                                st.download_button(
                                    label="⬇️ Download JSON File",
                                    data=json_string.encode('utf-8'),
                                    file_name=filename,
                                    mime="application/json",
                                    key=f"download_{file_hash}"
                                )
                            
                            with col2:
                                st.markdown("**Option 2: Copy Text**")
                                if st.button("📋 Show Copyable JSON", key=f"show_json_{file_hash}"):
                                    st.session_state[f"show_json_{file_hash}"] = True
                            
                            # Show copyable JSON if requested
                            if st.session_state.get(f"show_json_{file_hash}", False):
                                st.subheader("Copy for Postman")
                                st.info("💡 **Tip:** Select all text below (Ctrl+A) and copy (Ctrl+C)")
                                
                                # Use code block instead of text_area to prevent hanging
                                st.code(json_string, language='json')
                                
                                # Also provide a smaller text area for easier copying
                                st.text_area(
                                    "Copyable JSON (click in box and Ctrl+A to select all):",
                                    value=json_string,
                                    height=150,
                                    key=f"copy_area_{file_hash}"
                                )
                            
                            # Postman instructions
                            with st.expander("📋 Postman Setup Instructions"):
                                st.markdown("""
                                **How to use this JSON in Postman:**
                                
                                1. Open Postman and create a new POST request
                                2. Set the request URL to your API endpoint
                                3. Go to the **Body** tab
                                4. Select **raw** and choose **JSON** from the dropdown
                                5. Copy the JSON from the text area above
                                6. Paste it into the body field in Postman
                                7. Add any required headers (Content-Type: application/json, etc.)
                                8. Send the request
                                
                                **Common Headers:**
                                - `Content-Type: application/json`
                                - `Authorization: Bearer <your-token>` (if required)
                                """)
                        
                        # Statistics
                        st.header("4. Conversion Statistics")
                        col1, col2, col3, col4 = st.columns(4)
                        
                        with col1:
                            st.metric("Buildings Processed", len(api_json.get('buildings', [])))
                        
                        with col2:
                            services_count = sum(len(b.get('services', [])) for b in api_json.get('buildings', []))
                            st.metric("Total Services", services_count)
                        
                        with col3:
                            equipment_count = sum(len(s.get('equipment', [])) for b in api_json.get('buildings', []) for s in b.get('services', []))
                            st.metric("Equipment Items", equipment_count)
                        
                        with col4:
                            json_size = len(json.dumps(api_json))
                            st.metric("JSON Size", f"{json_size:,} chars")
                        
                    else:
                        st.error("❌ Failed to generate JSON from Excel data")
                        
                except Exception as e:
                    st.error(f"❌ Error generating JSON: {str(e)}")
                    with st.expander("Debug Information"):
                        st.code(traceback.format_exc())
            else:
                st.error("❌ Failed to load Excel file. Please check the file format and try again.")
    
    else:
        # Instructions when no file is uploaded
        st.info("👆 Please upload an Excel file to get started")
        
        with st.expander("📋 Excel File Requirements"):
            st.markdown("""
            **Your Excel file must contain:**
            
            - A sheet named **"Janitorial Services"**
            - Data starting from **row 4**
            - Specific columns in the expected positions (B, C, D, etc.)
            
            **Expected columns:**
            - Column B: Customer Name
            - Column C: Building ID
            - Column D: Address
            - Column I: Building Type
            - Column J: Total Square Footage
            - Column K: Cleanable Square Footage
            - Columns R-X: Weekly schedule (Sun-Sat)
            - And many more...
            
            **Multiple Buildings:**
            - Each row represents one building
            - All rows with the same customer will be grouped into one deal
            - Each building can have its own services, equipment, and schedules
            """)

if __name__ == "__main__":
    main()